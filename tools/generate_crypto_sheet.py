#!/usr/bin/env python3
"""
Generate styled Excel report from summarized RootData crypto fundraising data.

Columns: Date | Company (hyperlinked) | Info | Round | Amount ($) | Investors
Dark navy header, wrapped text, frozen top row.

Input:  .tmp/summarized_rootdata.json
Output: output/Crypto_News_YYYYMMDD_YYYYMMDD.xlsx
"""

import os
import sys
import json
import argparse
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

ROOTDATA_BASE = "https://www.rootdata.com"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
CELL_FONT   = Font(size=10, name="Calibri")
LINK_FONT   = Font(size=10, name="Calibri", color="0563C1", underline="single")
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

COLUMNS = [
    ("Date",        12),
    ("Company",     24),
    ("Info",        55),
    ("Round",       15),
    ("Amount ($)",  14),
    ("Investors",   35),
]


def generate_crypto_sheet(input_file: str, start_date: str, end_date: str,
                           output_dir: str) -> str:
    """
    Read summarized deals JSON and write a styled Excel file.

    Returns the path of the created file.
    """
    with open(input_file, encoding="utf-8") as f:
        deals = json.load(f)

    if not deals:
        raise RuntimeError(
            "No deals found — try expanding date range or lowering min_amount"
        )

    # Sort by date ascending
    def _sort_key(d):
        try:
            return datetime.strptime(d.get("date", ""), "%Y-%m-%d")
        except ValueError:
            return datetime.min

    deals = sorted(deals, key=_sort_key)

    # Build output filename
    s = start_date.replace("-", "")
    e = end_date.replace("-", "")
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, f"Crypto_News_{s}_{e}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Crypto Fundraising"

    # Header row
    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, deal in enumerate(deals, 2):
        company     = deal.get("company", "")
        company_url = deal.get("company_url", "")
        investors   = deal.get("investors", [])
        investors_str = ", ".join(investors) if isinstance(investors, list) else str(investors)

        row_data = [
            deal.get("date", ""),
            company,
            deal.get("info", ""),
            deal.get("round", ""),
            deal.get("amount_raw", ""),
            investors_str,
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER

        # Hyperlink company cell to RootData profile
        if company_url:
            company_cell = ws.cell(row=row_idx, column=2)
            full_url = company_url if company_url.startswith("http") else ROOTDATA_BASE + company_url
            company_cell.hyperlink = full_url
            company_cell.font = LINK_FONT

        ws.row_dimensions[row_idx].height = 60

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"✓ Saved {len(deals)} rows → {out_path}")
    return out_path


def main():
    parser = argparse.ArgumentParser(
        description="Generate styled Excel from summarized RootData crypto deals"
    )
    parser.add_argument("--input",      required=True, help="Summarized JSON input file")
    parser.add_argument("--start_date", required=True, help="Report start date YYYY-MM-DD")
    parser.add_argument("--end_date",   required=True, help="Report end date YYYY-MM-DD")
    parser.add_argument("--output_dir", default="output", help="Output directory")
    args = parser.parse_args()

    generate_crypto_sheet(
        input_file=args.input,
        start_date=args.start_date,
        end_date=args.end_date,
        output_dir=args.output_dir,
    )


if __name__ == "__main__":
    main()
